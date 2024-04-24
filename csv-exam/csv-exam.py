import sys

# pathの設定 (pip showで出てきた、LocationのPATHを以下に設定) Set path
sys.path.append(
    "/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/csv-examenv/lib/python3.12/site-packages"
)

import pandas as pd
import openpyxl
import os
import shutil

csv_path = "/Users/atoatoatomu/Downloads/exams.csv"

# create folder called trash, and it'll remove
os.mkdir("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/trash")

# CSVファイルの読み込み
data = pd.read_csv(csv_path)

# Excel形式で出力
data.to_excel("excel-data.xlsx")

# wb = openpyxl.load_workbook("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/excel-data.xlsx")
excel_data_wb = openpyxl.load_workbook("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/excel-data.xlsx")

def insert_col(wb, new_file_name):
    ws = wb["Sheet1"]
    ws.insert_cols(7)
    wb.save(new_file_name)

def rename_col(previous_file_name, new_file_name, new_col_name, col_number):
    df = pd.read_excel(previous_file_name)

    #▼列名の変更
    print({'Unnamed: ' + str(col_number) : new_col_name})
    df = df.rename(columns = {'Unnamed: ' + str(col_number) : new_col_name})

    #▼書き出し
    df.to_excel(new_file_name, index = None)


def xlsx_to_dict(ws, list_name):
    header_cells = None

    for row in ws.rows:
        if row[0].row == 1:
            # １行目
            header_cells = row
            # This code can have a name of a specific row
            # print(header_cells[1].value)
        else:
            # ２行目以降
            row_dic = {}
            # セルの値を「key-value」で登録
            # zip() is a built-in function that allows you to aggregate elements from multiple iterables into a single iterable
            # it loops together in parallel
            for k, v in zip(header_cells, row):
                row_dic[k.value] = v.value
            list_name.append(row_dic)

# task 1
def calculate_average():
    exam_list = []
    excel_data_ws = excel_data_wb["Sheet1"]

    # convert xlsx to dictionary
    xlsx_to_dict(excel_data_ws, exam_list)

    # add dict average values
    for exam in exam_list:
        average_score = (exam["math score"] + exam["reading score"] + exam["writing score"]) / 3
        exam["average score"] = int(average_score)
        print(int(average_score))

    # insert col
    insert_col(excel_data_wb, "trash/after-insert-average-col.xlsx")

    # rename average col
    # number 6 will shift to 7
    rename_col("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/trash/after-insert-average-col.xlsx", "trash/after-rename-and-insert-average.xlsx", "Average of points", 6)

    average_wb = openpyxl.load_workbook("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/trash/after-rename-and-insert-average.xlsx")
    sheet = average_wb.active

    # add values to average excel file renamed
    for i in range(len(exam_list)):
        sheet_index = i + 2
        sheet["G" + str(sheet_index)] = exam_list[i]["average score"]

    average_wb.save('result-calculate-average.xlsx')

calculate_average()

# additional task 1
def calculate_total(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"]

    insert_col(wb, "trash/after-insert-sum-col.xlsx")
    rename_col("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/trash/after-insert-sum-col.xlsx", "trash/after-rename-and-insert-sum.xlsx", "Sum", 6)

    exam_list = []

    xlsx_to_dict(ws, exam_list)

    for exam in exam_list:
        total_score = exam["math score"] + exam["reading score"] + exam["writing score"]
        exam["total_score"] = total_score

    # for manipulating new sheet (after inserting and renaming sum col)
    sum_wb = openpyxl.load_workbook("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/trash/after-rename-and-insert-sum.xlsx")
    sheet = sum_wb.active

    # add sum value into G column
    for i in range(len(exam_list)):
        sheet_index = i + 2
        sheet['G' + str(sheet_index)] = exam_list[i]["total_score"]
    sum_wb.save('result-calculate-sum.xlsx')

calculate_total("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/result-calculate-average.xlsx")

# remove trash folder
shutil.rmtree("/Users/atoatoatomu/Desktop/2024-programming-class/python_review/csv-exam/trash")